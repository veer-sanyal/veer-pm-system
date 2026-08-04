#!/usr/bin/env python3
"""Turn the internal Sagamore promo calendar into the copy Bryon actually receives.

The internal workbook talks ABOUT the client in the third person ("Bryon's grid", "Confirm with
Bryon", "ask Bryon if blue means tentative") and carries version-changelog language ("v0.3 - LAYER
ADDED", "New 'Observances' tab"). Both read badly when the client is the one opening the file.

Run this after regenerating the workbook. It rewrites those cells to second person, drops the
version scaffolding, and writes a separate output file so the internal copy stays intact.

    python3 tools/make-client-copy.py \\
        ~/Downloads/"Sagamore Promotional Calendar DRAFT v0.3.xlsx" \\
        ~/Downloads/"Sagamore Council Promotional Calendar - Draft.xlsx"
"""
import re
import sys

import openpyxl

CELLS = {
    "Read Me": {
        "A1": "Sagamore Council Promotional Calendar - Draft",
        "A2": "Prepared by the Purdue Daniels externship team | Built from your Communication Grid "
              "and your 90 to 120 day promotion rule",
        "A5": "Your Communication Grid says when events HAPPEN. This calendar says when to TALK "
              "about each one, working backward from each event:",
        "A14": "EDITING",
        "A17": "ASSUMPTIONS AND OPEN QUESTIONS FOR YOU",
        "A18": "- The 90 to 120 day rule is taken from your Jul 20 email; modeled as: announce 3 "
               "months before, details 2 months, weekly ramp 1 month.",
        "A19": "- The December mark on the 'University of Scouting (spring, IU Kokomo)' row was "
               "moved to the fall/Purdue row (Dec 4-6). Please confirm.",
        "A20": "- Date differences between the grid and the website, to confirm: Trade-o-Ree (grid "
               "Sep, 2026 actual Aug 21-22); fall Cub Scout Shoot Out (grid Oct, 2026 actual Sep 26).",
        "A21": "- Two blue cells in your grid (Pinewood Derby Feb, Fall Camporee Nov). Does blue "
               "mean tentative or new?",
        "A22": "- Eagle Scout Reception (Dec 4) was added from the website; it was not a row in "
               "your grid.",
        "A33": "HOLIDAYS AND AWARENESS OBSERVANCES",
        "A34": "The 'Observances' tab lists holidays and awareness months, each with the audience "
               "it serves and how to use it.",
        "A35": "The 'Posting Schedule' tab flags any post falling within 3 days of an observance, "
               "plus the month-long ones.",
        "A36": "The observance list is a MENU, not a commitment. The Status column flags the ones "
               "that need your call.",
        "A38": "OPEN FOR YOU: (1) Pride Month, your decision rather than ours; (2) the council's "
               "actual Scouting for Food dates;",
    },
    "Promo Calendar": {
        "Q8": "Blue cell in your grid. Please confirm status.",
        "Q18": "Blue cell in your grid. Please confirm status.",
        "Q20": "Per your grid, the only item currently on a calendar",
        "Q21": "Grid says Oct; website says Sep 26. Please confirm.",
        "Q25": "Grid says Sep; 2026 actual is Aug. Please confirm.",
        "Q28": "The Dec mark sat on the IU Kokomo row in the grid and was moved here. Please confirm.",
        "Q29": "Added from the website; not a row in your grid",
    },
    "Observances": {
        "E12": "Your decision. Nothing is scheduled here without your explicit go-ahead.",
        "E21": "Scouting for Food drives run around here. Please confirm the council's actual dates.",
    },
}

# Status column on the Observances tab, client-facing vocabulary.
STATUS = {"Confirm with Bryon": "Needs your input", "Bryon decides": "Your decision"}

# Anything matching this in the output is a leak. The attribution line in A2 is the one allowed hit.
LEAK = re.compile(r"bryon|\bv0\.\d", re.I)


def main(src, out):
    wb = openpyxl.load_workbook(src)

    for sheet, cells in CELLS.items():
        ws = wb[sheet]
        for ref, value in cells.items():
            ws[ref] = value

    ws = wb["Observances"]
    for (cell,) in ws.iter_rows(min_col=6, max_col=6):
        if cell.value in STATUS:
            cell.value = STATUS[cell.value]

    wb.save(out)

    leaks = [
        f"{w.title}!{c.coordinate}: {c.value}"
        for w in openpyxl.load_workbook(out).worksheets
        for row in w.iter_rows()
        for c in row
        if isinstance(c.value, str) and LEAK.search(c.value)
    ]
    if leaks:
        print("THIRD-PERSON LEAK, do not send:", *leaks, sep="\n  ")
        return 1
    print(f"wrote {out}")
    return 0


def demo():
    """Self-check: the substitutions cover every third-person reference we know about."""
    known = [
        "Bryon's grid says when events HAPPEN.",
        "ASSUMPTIONS + OPEN QUESTIONS FOR BRYON",
        "Blue cell in Bryon's grid - confirm status",
        "Confirm with Bryon",
        "Bryon decides",
        "v0.3 - HOLIDAY + AWARENESS-OBSERVANCE LAYER ADDED",
    ]
    replacements = [v for cells in CELLS.values() for v in cells.values()] + list(STATUS.values())
    for original in known:
        assert LEAK.search(original), f"detector misses: {original}"
    for replacement in replacements:
        assert not LEAK.search(replacement) or replacement.startswith("Prepared by"), \
            f"replacement still leaks: {replacement}"
    print("demo ok")


if __name__ == "__main__":
    if len(sys.argv) == 2 and sys.argv[1] == "--demo":
        demo()
    elif len(sys.argv) == 3:
        sys.exit(main(sys.argv[1], sys.argv[2]))
    else:
        sys.exit(__doc__)
