#!/usr/bin/env python3
"""Add the holiday / awareness-observance layer to the Sagamore promo calendar.

Reads v0.2, writes v0.3. Does not modify the input (Excel may have it open).
The observance list is a MENU for Bryon to approve, not a commitment - see the
`Status` column on the Observances sheet.
"""
import sys
from datetime import date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SRC = Path.home() / "Downloads" / "Sagamore Promotional Calendar DRAFT v0.2.xlsx"
DST = Path.home() / "Downloads" / "Sagamore Promotional Calendar DRAFT v0.3.xlsx"

# ponytail: a post is "near" an observance if it falls within this window either side
NEAR_DAYS = 3


def nth_weekday(year, month, weekday, n):
    """n-th `weekday` (Mon=0) of month. n=-1 means last."""
    if n > 0:
        d = date(year, month, 1)
        d += timedelta(days=(weekday - d.weekday()) % 7)
        return d + timedelta(weeks=n - 1)
    d = date(year, month, 28)
    while (d + timedelta(days=1)).month == month:
        d += timedelta(days=1)
    return d - timedelta(days=(d.weekday() - weekday) % 7)


def scout_sunday(year):
    """Traditionally the Sunday preceding Feb 8 (Scouting America's founding)."""
    d = date(year, 2, 8)
    return d - timedelta(days=(d.weekday() + 1) % 7 or 7)


# (name, audience, hook, status) for month-long observances, keyed by month
MONTHLY = {
    1: [],
    2: [("Black History Month", "General public",
         "Profile Black Scouts, leaders and troops in the council; archive photos work well.",
         "Confirm with Bryon")],
    3: [("Women's History Month", "Prospective parents / volunteers",
         "Spotlight women den leaders and volunteers - directly serves the ~24-26 parent audience.",
         "Confirm with Bryon")],
    4: [("Volunteer Appreciation (National Volunteer Month)", "Volunteer leaders",
         "A month of volunteer spotlights; pairs with the Council Annual Luncheon recap.",
         "Recommended")],
    5: [("Asian American & Pacific Islander Heritage Month", "General public",
         "Same treatment as February and March if the council has stories to tell.",
         "Confirm with Bryon"),
        ("Jewish American Heritage Month", "General public",
         "Only if there is a real story; skip rather than post a stock graphic.",
         "Optional")],
    6: [("Pride Month", "General public",
         "CLIENT DECISION - do not schedule without Bryon's explicit call.",
         "Bryon decides")],
    7: [],
    8: [],
    9: [("Hispanic Heritage Month (Sep 15 - Oct 15)", "General public",
         "Spans two months; one post in each half.", "Confirm with Bryon")],
    10: [("Hispanic Heritage Month (Sep 15 - Oct 15)", "General public",
          "Second half of the observance.", "Confirm with Bryon")],
    11: [("Native American Heritage Month", "General public",
          "Scouting has real program ties here; handle with care and local voices.",
          "Confirm with Bryon")],
    12: [],
}


def fixed_days(year):
    """(date, name, audience, hook, status) single-day observances."""
    return [
        (nth_weekday(year, 1, 0, 3), "MLK Jr. Day of Service", "General public",
         "Service is the easiest Scouting story to tell. Pair with a unit service project.",
         "Recommended"),
        (scout_sunday(year), "Scout Sunday / Scout Sabbath", "Scout families / volunteers",
         "Units in uniform at their chartered organizations - photo-rich, zero design needed.",
         "Recommended"),
        (date(year, 2, 8), "Scouting America anniversary + National Eagle Scout Day",
         "All audiences",
         "The single best storytelling day of the year. Eagle Scout spotlights, founding history.",
         "Recommended"),
        (nth_weekday(year, 2, 0, 3), "Presidents Day", "General public",
         "Low priority; only if a unit does something civic.", "Optional"),
        (nth_weekday(year, 5, 0, -1), "Memorial Day", "General public / donors",
         "Flag placement at cemeteries is a signature Scouting activity. Shoot photos.",
         "Recommended"),
        (date(year, 6, 14), "Flag Day", "General public",
         "Flag retirement ceremonies - visually strong, unique to Scouting.", "Recommended"),
        (date(year, 7, 4), "Independence Day", "General public",
         "Parade participation; recap post the same weekend.", "Recommended"),
        (date(year, 9, 11), "Patriot Day", "General public",
         "Somber tone, no promotion. One respectful post or skip.", "Optional"),
        (date(year, 9, 17), "Constitution Day / Constitution Week", "Volunteer leaders",
         "Ties to the Citizenship merit badges.", "Optional"),
        (date(year, 11, 11), "Veterans Day", "General public / donors",
         "Flag ceremonies and veteran volunteers. Strong donor-audience content.",
         "Recommended"),
        (nth_weekday(year, 11, 3, 4), "Thanksgiving", "Scout families",
         "Scouting for Food drives run around here - confirm the council's actual dates.",
         "Confirm with Bryon"),
        (nth_weekday(year, 11, 3, 4) + timedelta(days=5), "GivingTuesday",
         "Donors / business community",
         "The one date on this list that asks for money directly. Plan it, do not improvise.",
         "Recommended"),
        (date(year, 12, 25), "Christmas / holiday close", "All audiences",
         "Year-in-review photo dump; thank volunteers and donors by name.", "Recommended"),
    ]


def main():
    if not SRC.exists():
        sys.exit(f"missing: {SRC}")
    wb = openpyxl.load_workbook(SRC)
    sched = wb["Posting Schedule"]

    # --- build the lookup ---
    years = {2026, 2027}
    single = {}
    for y in years:
        for d, name, aud, hook, status in fixed_days(y):
            single.setdefault(d, []).append((name, aud, hook, status))

    # --- new column on Posting Schedule ---
    col = sched.max_column + 1
    hdr = sched.cell(row=1, column=col, value="Observance / seasonal hook")
    hdr.font = Font(bold=True)
    tagged = 0
    monthly_used = set()  # one monthly-observance nod per month, on its first post
    for r in range(2, sched.max_row + 1):
        raw = sched.cell(row=r, column=1).value
        if raw is None:
            continue
        d = raw.date() if hasattr(raw, "date") else _parse(str(raw))
        if d is None:
            continue
        hits = []
        for off in range(-NEAR_DAYS, NEAR_DAYS + 1):
            for name, _aud, _hook, status in single.get(d + timedelta(days=off), []):
                hits.append(f"{name} ({(d + timedelta(days=off)).strftime('%b %d')})")
        if (d.year, d.month) not in monthly_used and MONTHLY.get(d.month):
            monthly_used.add((d.year, d.month))
            hits += [name for name, *_ in MONTHLY[d.month]]
        if hits:
            sched.cell(row=r, column=col, value="; ".join(dict.fromkeys(hits)))
            tagged += 1
    sched.column_dimensions[hdr.column_letter].width = 46

    # --- Observances sheet ---
    if "Observances" in wb.sheetnames:
        del wb["Observances"]
    ws = wb.create_sheet("Observances", wb.sheetnames.index("Posting Schedule"))
    head = ["When", "Observance", "Type", "Primary audience", "How to use it", "Status"]
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    for m in range(1, 13):
        label = date(2026, m, 1).strftime("%B")
        for name, aud, hook, status in MONTHLY[m]:
            ws.append([label, name, "Month-long", aud, hook, status])
        for d, name, aud, hook, status in sorted(fixed_days(2027)):
            if d.month == m:
                ws.append([f"{label} {d.day} (2027)", name, "Single day", aud, hook, status])
    widths = [18, 46, 12, 32, 78, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

    # --- Read Me note ---
    rm = wb["Read Me"]
    for line in [
        "",
        "v0.3 - HOLIDAY + AWARENESS-OBSERVANCE LAYER ADDED",
        "New 'Observances' tab: holidays and awareness months, each with the audience it serves and how to use it.",
        "New column on 'Posting Schedule': flags any post falling within 3 days of an observance, plus the month-long ones.",
        "The observance list is a MENU, not a commitment - the Status column says which need Bryon's call.",
        "Rule of thumb: an observance only earns a post if the council has a REAL photo or story for it. Skip over stock graphics.",
        "OPEN FOR BRYON: (1) Pride Month - client decision, not ours; (2) the council's actual Scouting for Food dates;",
        "  (3) which heritage months the council has real stories for, so we schedule those and drop the rest.",
    ]:
        rm.cell(row=rm.max_row + 1, column=1, value=line)

    wb.save(DST)
    print(f"wrote {DST}\n  tagged {tagged} of {sched.max_row - 1} scheduled posts")


def _parse(s):
    from datetime import datetime
    for fmt in ("%b %d, %Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


if __name__ == "__main__":
    main()
